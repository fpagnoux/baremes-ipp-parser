from openfisca_core.parameters import ParameterNode
import openpyxl

wb = openpyxl.load_workbook('/Users/florianpagnoux/dev/openfisca/baremes-ipp/baremes-ipp-prestations-sociales-social-benefits.xlsx')

sheet = wb['AF_CM']

def clean_none_values(param_data):
  values = param_data['values']
  sorted_dates = sorted(values.keys())
  first_value_found = False
  first_none_found = False
  for date in sorted_dates:
    value = values[date]['value']
    if value is None and (first_none_found or not first_value_found):
      del values[date]
    elif value is None:
      first_none_found = True
    elif value is not None and not first_value_found:
      first_value_found = True


class SheetParser(object):

  def __init__(self, sheet):
    self.sheet = sheet
    self.date_column = None
    self.reference_column = None
    self.data_columns = []
    self.dates = None
    self.references = None
    self.first_data_row = None
    self.last_data_row = None

  # def read_value(self, cell):
  #   """
  #   Read the value of a cell, including if the cell has been merged horizontally
  #   """
  #   if cell.internal_value is not None:
  #     return cell.internal_value
  #   for cell_range in sheet.merged_cells.ranges:
  #     if cell_range.min_row <= cell.row <= cell_range.max_row and cell_range.min_col <= cell.col_idx <= cell_range.max_col:
  #       return self.sheet.cell(cell_range.min_row, cell_range.min_col).internal_value

  def unmerge_cells(self):
    merged_ranges = self.sheet.merged_cells.ranges
    for cell_range in merged_ranges:
      sheet.unmerge_cells(cell_range.coord)
      main_cell = self.sheet.cell(cell_range.min_row, cell_range.min_col)
      for column in range(cell_range.min_col + 1, cell_range.max_col + 1):
        cell = self.sheet.cell(cell_range.min_row, column)
        cell.set_explicit_value(main_cell.internal_value)

  def parse_headers(self):
    for cell in self.sheet['1']:
      key = cell.internal_value
      if key is None:
        break
      elif key == 'date':
        self.date_column = cell.column
      elif key == 'reference':
        self.reference_column = cell.column
      elif key == 'date_parution_jo' or key == 'notes':
        pass # Ignore those columns for the moment
      else:
        self.data_columns.append(cell.column)

  def parse_dates(self):
    dates = []
    visited_a_date = False
    for cell in self.sheet[self.date_column][2:]:
      if cell.internal_value is None:
        if not visited_a_date: # We are still in the header
          continue
        # Once you reach a blank cell in the date column, stop
        else:
          self.last_data_row = cell.row - 1
          break
      if not visited_a_date:
        visited_a_date = True
        self.first_data_row = cell.row
      date = cell.internal_value .strftime('%Y-%m-%d')
      dates.append(date)
    self.dates = dates
    self.number_values = len(self.dates)

  def parse_references(self):
    references = []
    for cell in self.sheet[self.reference_column][self.first_data_row - 1:self.last_data_row]:
      references.append(cell.internal_value)
    self.references = references

  def build_description(self, column):
    description_cells = column[1:self.first_data_row -1]
    return "; ".join([
      cell.internal_value
      for cell in description_cells
      if cell.internal_value is not None
      ])

  def parse_column(self, column):
    data = {}
    code = column[0].internal_value
    data = { 'description': self.build_description(column), 'values': {} }
    for date, reference, cell in zip(self.dates, self.references, column[self.first_data_row - 1:self.last_data_row]):
      item = {'value': cell.internal_value}
      if reference is not None:
        item['reference'] = reference
      data['values'][date] = item

    clean_none_values(data)

    return code, data

  def parse(self):
    self.unmerge_cells()
    self.parse_headers()
    self.parse_dates()
    self.parse_references()

    parsed_columns = [self.parse_column(self.sheet[column_name]) for column_name in self.data_columns]

    return ParameterNode('', data = {
      code: data for code, data in parsed_columns
    })


parser = SheetParser(sheet)
x = parser.parse()
